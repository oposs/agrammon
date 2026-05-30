#!/usr/bin/env python3
"""Validate a generated .xlsx: XML well-formedness + content/format round-trip."""
import sys, zipfile, xml.dom.minidom as M
import openpyxl

fn = sys.argv[1] if len(sys.argv) > 1 else "sample-fast.xlsx"

z = zipfile.ZipFile(fn)
ok = bad = 0
for n in z.namelist():
    try:
        M.parseString(z.read(n)); ok += 1
    except Exception as e:
        bad += 1; print("MALFORMED", n, e)
print(f"{fn}: {ok}/{len(z.namelist())} parts well-formed")

wb = openpyxl.load_workbook(fn)
print("sheets:", wb.sheetnames)
ws = wb["Eingaben"]
print("Eingaben row6:", [ws.cell(6, c).value for c in range(1, 6)])
# D6 (col index 4) is the numeric value, written with num-format 0.000 + right align
print("  D6 value:", ws.cell(6, 4).value,
      "| numfmt:", ws.cell(6, 4).number_format,
      "| align:", ws.cell(6, 4).alignment.horizontal)
print("  colA width:", ws.column_dimensions["A"].width,
      "| colD width:", ws.column_dimensions["D"].width)
wif = wb["Eingaben formatiert"]
print("EinFmt A6:", repr(wif["A6"].value), "| bold:", wif["A6"].font.bold)
erf = wb["Ergebnisse formatiert"]
print("ErgFmt A1:", repr(erf["A1"].value), "| bold:", erf["A1"].font.bold)
sys.exit(1 if bad else 0)
